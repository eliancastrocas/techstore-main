from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from .models import Product, Service, FormRequest, StockMovement, ProductDetails
from .forms import ProductForm, ServiceForm, FormRequestForm
from users.models import UserProfile
from django.db.models import Q
from django.http import HttpResponse, FileResponse

from django.conf import settings
import os
import csv
from .dian_api import get_cities, get_departments

from cart.models import Cart

import logging
logger = logging.getLogger(__name__)

def _get_vendor_profile_or_none(user):
    """Devuelve el perfil del vendedor autenticado si existe, si no None."""
    if not getattr(user, 'is_authenticated', False):
        return None
    profile = getattr(user, 'profile', None)
    if profile is None:
        return None
    # Normalizar validación de rol (evitar mezclar role vs is_vendedor)
    if hasattr(profile, 'is_vendedor'):
        if profile.is_vendedor():
            return profile
        return None
    if getattr(profile, 'role', None) == 'vendedor':
        return profile
    return None

def _debug_permission_denied(request, product=None, reason=''):
    vendor_profile = _get_vendor_profile_or_none(request.user)
    user_repr = f"{request.user.id}:{request.user.username}" if request.user.is_authenticated else 'anon'
    payload = {
        'user': user_repr,
        'profile_id': getattr(vendor_profile, 'id', None),
        'product_id': getattr(product, 'id', None),
        'product_seller_id': getattr(product, 'seller_id', None),
        'reason': reason,
    }
    # Mensajes visibles para depuración (temporal)
    messages.error(request, f"DEBUG permiso denegado: {payload}")
    logger.warning('Permission denied: %s', payload)



def product_search(request):
    query = request.GET.get("q", "").strip()
    category = request.GET.get("category", "").strip()
    min_price = request.GET.get("min_price", "").strip()
    max_price = request.GET.get("max_price", "").strip()
    in_stock = request.GET.get("in_stock", "").strip()

    # Search only vendor's products if logged in as vendor
    if request.user.is_authenticated and hasattr(request.user, 'profile') and request.user.profile.is_vendedor():
        products = Product.objects.filter(seller=request.user.profile).order_by("-created_at")
    else:
        # Clients and guests see available products
        products = Product.objects.filter(stock__gt=0).order_by("-created_at")

    if query:
        products = products.filter(
            Q(name__icontains=query) | Q(description__icontains=query)
        )

    if category:
        products = products.filter(category=category)

    if min_price:
        try:
            products = products.filter(price__gte=float(min_price))
        except ValueError:
            pass

    if max_price:
        try:
            products = products.filter(price__lte=float(max_price))
        except ValueError:
            pass

    if in_stock:
        products = products.filter(stock__gt=0)

    cart = Cart.get_cart(request)
    return render(
        request,
        "products/product_list.html",
        {
            "products": products,
            "query": query,
            "category": category,
            "min_price": min_price,
            "max_price": max_price,
            "in_stock": in_stock,
            "cart": cart,
            "cart_count": cart.item_count,
        },
    )


def product_catalog(request):
    """Public product catalog for clients and guests"""
    query = request.GET.get("q", "").strip()
    category = request.GET.get("category", "").strip()
    min_price = request.GET.get("min_price", "").strip()
    max_price = request.GET.get("max_price", "").strip()

    products = Product.objects.filter(stock__gt=0).order_by("-created_at")

    if query:
        products = products.filter(
            Q(name__icontains=query) | Q(description__icontains=query)
        )

    if category:
        products = products.filter(category=category)

    if min_price:
        try:
            products = products.filter(price__gte=float(min_price))
        except ValueError:
            pass

    if max_price:
        try:
            products = products.filter(price__lte=float(max_price))
        except ValueError:
            pass

    cart = Cart.get_cart(request)
    return render(
        request,
        "products/product_catalog.html",
        {
            "products": products,
            "query": query,
            "category": category,
            "min_price": min_price,
            "max_price": max_price,
            "cart": cart,
            "cart_count": cart.item_count,
        },
    )


@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and u.profile.role == "vendedor")
def product_list(request):
    profile = request.user.profile
    products = Product.objects.filter(seller=profile).order_by("-created_at")

    # Get filter parameters from GET request
    query = request.GET.get("q", "").strip()
    category = request.GET.get("category", "").strip()
    min_price = request.GET.get("min_price", "").strip()
    max_price = request.GET.get("max_price", "").strip()
    in_stock = request.GET.get("in_stock", "").strip()

    if query:
        products = products.filter(
            Q(name__icontains=query) | Q(description__icontains=query)
        )

    if category:
        products = products.filter(category=category)

    if min_price:
        try:
            products = products.filter(price__gte=float(min_price))
        except ValueError:
            pass

    if max_price:
        try:
            products = products.filter(price__lte=float(max_price))
        except ValueError:
            pass

    if in_stock:
        products = products.filter(stock__gt=0)

    cart = Cart.get_cart(request)
    return render(request, "products/product_list.html", {
        "products": products,
        "query": query,
        "category": category,
        "min_price": min_price,
        "max_price": max_price,
        "in_stock": in_stock,
        "cart": cart,
        "cart_count": cart.item_count,
    })





from django.contrib.auth.decorators import user_passes_test


@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and getattr(u.profile, "role", None) == "vendedor")
def product_create(request):
    try:
        profile = request.user.profile
    except UserProfile.DoesNotExist:
        profile = None

    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product_name = form.cleaned_data.get("name", "").strip()
            category = form.cleaned_data.get("category", "otro")
            price = form.cleaned_data.get("price", 0)
            description = form.cleaned_data.get("description", "")
            image_url = form.cleaned_data.get("image_url", "")
            stock = form.cleaned_data.get("stock", 0)
            is_damaged = form.cleaned_data.get("is_damaged", False)

            product, created = Product.objects.update_or_create(
                name=product_name,
                seller=profile,
                defaults={
                    "category": category,
                    "price": price,
                    "description": description,
                    "image_url": image_url,
                    "stock": stock,
                    "is_damaged": is_damaged,
                },
            )

            # ProductDetails se guarda dentro de ProductForm.save()
            # (si el form se usa correctamente). Como aquí usamos update_or_create
            # con campos manuales, aseguramos la existencia de details.
            if not hasattr(product, "details"):
                ProductDetails.objects.create(product=product)

            # Guardar detalles adicionales desde el POST
            details_fields = [
                "referencias",
                "especificaciones_tecnicas",
                "caracteristicas",
                "contenido_caja",
                "compatibilidades",
                "dimensiones",
                "materiales",
                "garantia_detalle",
                "otras_comentarios",
            ]
            details, _ = ProductDetails.objects.get_or_create(product=product)
            for f in details_fields:
                setattr(details, f, request.POST.get(f, ""))
            details.save()

            if created:
                # Only create movimento if there's NO existing entrada for this product
                if product.stock > 0 and profile:
                    StockMovement.objects.create(
                        product=product,
                        movement_type="entrada",
                        quantity=product.stock,
                        reason="Producto criado",
                        created_by=profile,
                    )
                messages.success(request, "Producto creado exitosamente.")
            else:
                messages.success(request, "Producto actualizado exitosamente.")
            return redirect("product_list")
    else:
        form = ProductForm()
    return render(
        request, "products/product_form.html", {"form": form, "action": "Criar"}
    )


@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and u.profile.is_vendedor())
def product_update(request, pk):
    vendor_profile = _get_vendor_profile_or_none(request.user)
    if not vendor_profile:
        _debug_permission_denied(request, product=None, reason='usuario sin perfil vendedor')
        return redirect('product_list')

    # Cargar únicamente el producto que pertenece al vendedor autenticado.
    # Si no existe, denegamos (evita validaciones “a posteriori”).
    try:
        product = Product.objects.select_related('seller__user').get(pk=pk, seller=vendor_profile)
    except Product.DoesNotExist:
        # Debug detallado de por qué falla.
        # Intentamos recuperar sin el filtro para poder mostrar seller_id (si existe).
        try:
            any_product = Product.objects.only('id', 'seller_id').get(pk=pk)
        except Product.DoesNotExist:
            any_product = None
        _debug_permission_denied(
            request,
            product=any_product,
            reason=f'no existe Product(pk={pk}, seller_id={vendor_profile.id})',
        )
        return redirect('product_list')

    if request.method == "POST":
        messages.info(
            request,
            f"DEBUG update OK: user={request.user.username}, vendor_profile_id={vendor_profile.id}, product_id={product.id}, product_seller_id={product.seller_id}",
        )
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, "Producto actualizado exitosamente.")
            return redirect("product_list")
    else:
        form = ProductForm(instance=product)

    return render(
        request,
        "products/product_form.html",
        {"form": form, "product": product, "action": "Editar"},
    )




@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and u.profile.is_vendedor())
def product_delete(request, pk):
    vendor_profile = _get_vendor_profile_or_none(request.user)
    if not vendor_profile:
        _debug_permission_denied(request, product=None, reason='usuario sin perfil vendedor')
        return redirect('product_list')

    # Cargar únicamente el producto que pertenece al vendedor autenticado.
    # Si no existe, denegamos.
    try:
        product = Product.objects.select_related('seller__user').get(pk=pk, seller=vendor_profile)
    except Product.DoesNotExist:
        # Debug detallado de por qué falla.
        try:
            any_product = Product.objects.only('id', 'seller_id').get(pk=pk)
        except Product.DoesNotExist:
            any_product = None
        _debug_permission_denied(
            request,
            product=any_product,
            reason=f'no existe Product(pk={pk}, seller_id={vendor_profile.id})',
        )
        return redirect('product_list')

    if request.method == "POST":
        messages.info(
            request,
            f"DEBUG delete OK: user={request.user.username}, vendor_profile_id={vendor_profile.id}, product_id={product.id}, product_seller_id={product.seller_id}",
        )
        product.delete()
        messages.success(request, "Producto eliminado exitosamente.")
        return redirect("product_list")

    return render(request, "products/product_confirm_delete.html", {"product": product})






def product_details(request, pk):
    product = get_object_or_404(Product, pk=pk)
    # Mostramos solo si hay stock > 0 (misma lógica de catálogo)
    if product.stock <= 0:
        messages.error(request, "Este producto no está disponible en este momento.")
        return redirect("product_catalog")

    # Garantizar que exista details
    try:
        details = product.details
    except ProductDetails.DoesNotExist:
        details = None

    from cart.models import Cart
    cart = Cart.get_cart(request)

    return render(
        request,
        "products/product_details.html",
        {
            "product": product,
            "details": details,
            "cart": cart,
            "cart_count": getattr(cart, "item_count", 0),
        },
    )


def services_catalog(request):
    """Página de servicios"""
    from cart.models import Cart
    from .models import Service

    cart = Cart.get_cart(request)
    services = Service.objects.all()
    return render(
        request,
        "products/services.html",
        {"cart": cart, "cart_count": cart.item_count, "services": services},
    )


@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and getattr(u.profile, "role", None) == "vendedor")
def vendor_form_requests(request):

    """Vendor dashboard for service requests sent by clients (reparación y mantenimiento)."""
    profile = request.user.profile

    # Para mantener compatibilidad, priorizamos las solicitudes asignadas al vendedor.
    # Según el requerimiento, clientes no deben acceder a esta vista.
    form_requests_qs = FormRequest.objects.filter(seller=profile)

    # Filtrar: inicialmente reparaciones y mantenimientos.
    # - Reparación: service_option='reparacion'
    # - Mantenimiento: service_option='mantenimiento'
    # Nota: no se importa "models" en este archivo; usar Q directamente.
    form_requests_qs = form_requests_qs.filter(
        Q(service_option="reparacion") | Q(service_option="mantenimiento")
    )


    # Mostrar primero pendientes (y luego el resto), dejando orden por fecha.
    form_requests = form_requests_qs.order_by("status", "-created_at")


    return render(
        request,
        "products/vendor_services_panel.html",
        {
            "form_requests": form_requests,
        },
    )



@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and getattr(u.profile, "role", None) == "vendedor")
def bulk_upload(request):
    profile = request.user.profile
    new_products = []

    if request.method == "POST":
        file = request.FILES.get("file")
        if file:
            ext = file.name.split(".")[-1].lower()
            temp_path = os.path.join(settings.MEDIA_ROOT, "bulk", file.name)
            os.makedirs(os.path.dirname(temp_path), exist_ok=True)

            with open(temp_path, "wb") as f:
                for chunk in file.chunks():
                    f.write(chunk)

            rows = []
            if ext == "csv":
                try:
                    with open(temp_path, "r", encoding="utf-8") as f:
                        reader = csv.DictReader(f)
                        rows = list(reader)
                except Exception as e:
                    messages.error(request, f"Error: {str(e)}")
            elif ext in ["xlsx", "xls"]:
                try:
                    from openpyxl import load_workbook

                    wb = load_workbook(temp_path, data_only=True)
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        rows.append(dict(zip(headers, row)))
                except ImportError:
                    messages.error(request, "Usa solo CSV. Excel no disponible.")
                except Exception as e:
                    messages.error(request, f"Error Excel: {str(e)}")
            else:
                messages.error(request, "Usa CSV o Excel.")

            if not rows:
                messages.error(request, "El archivo está vacío.")
            else:
                required = ["name", "price", "description", "stock"]
                first = rows[0] if rows else {}
                missing = [c for c in required if c not in first]
                if missing:
                    messages.error(request, f"Columnas faltantes: {', '.join(missing)}")
                else:
                    for row in rows:
                        name = str(row.get("name", "")).strip()
                        price = float(row.get("price", 0) or 0)
                        description = str(row.get("description", "")).strip()
                        stock = int(row.get("stock", 0) or 0)
                        category = str(row.get("category", "otro")).strip().lower()
                        image_url = str(row.get("image_url", "")).strip()

                        if name and price > 0:
                            stock_qty = abs(stock)
                            product, created = Product.objects.update_or_create(
                                name=name,
                                seller=profile,
                                defaults={
                                    "price": abs(price),
                                    "description": description,
                                    "stock": stock_qty,
                                    "category": category if category else "otro",
                                    "image_url": image_url if image_url else "",
                                },
                            )
                            if stock_qty > 0:
                                StockMovement.objects.create(
                                    product=product,
                                    movement_type="entrada",
                                    quantity=stock_qty,
                                    reason="Carga masiva",
                                    created_by=profile,
                                )
                            if created:
                                new_products.append(name)
                            else:
                                new_products.append(f"{name} (actualizado)")

                    new_count = sum(
                        1 for p in new_products if not p.endswith("(actualizado)")
                    )
                    updated_count = len(new_products) - new_count
                    msg = f"{new_count} productos importados."
                    if updated_count > 0:
                        msg += f" {updated_count} actualizados."
                    messages.success(request, msg)

    return render(request, "products/bulk_upload.html", {"new_products": new_products})


def cities_list(request):
    """Lista de ciudades colombianas desde API externa (con búsqueda)"""
    query = request.GET.get("q", "").strip().lower()
    
    all_cities = get_cities()
    all_departments = get_departments()
    dept_map = {d["id"]: d["name"] for d in all_departments}
    
    # Add dept_name to each city
    for city in all_cities:
        city["dept_name"] = dept_map.get(city.get("departmentId"), "Desconocido")
    
    if query:
        cities = [city for city in all_cities if query in city["name"].lower() or query in city["dept_name"].lower()]
    else:
        cities = all_cities
    
    context = {
        "cities": cities,
        "query": request.GET.get("q", ""),
        "count": len(cities),
        "total": len(all_cities),
    }
    return render(request, "products/cities.html", context)




@login_required
def reparacion_form(request):
    """Formulario específico para reparación de dispositivos"""
    from cart.models import Cart

    cart = Cart.get_cart(request)

    if request.method == "POST":
        form = FormRequestForm(request.POST, request.FILES)

        if form.is_valid():
            form_request = form.save(commit=False)





            # Asociar automáticamente al vendedor correcto.
            # Regla: el panel lista por FormRequest.seller, así que NO debe apuntar al cliente.
            # Asignamos el primer vendedor que esté verificado/activo; si no existe, mantenemos el valor actual.
            vendedor = None
            if hasattr(request.user, "profile"):
                try:
                    vendedor = UserProfile.objects.filter(role="vendedor").order_by("id").first()
                except Exception:
                    vendedor = None
            if vendedor:
                form_request.seller = vendedor



            form_request.customer_name = (
                request.user.get_full_name() or request.user.username
            )
            form_request.email = request.user.email or ''

            # Guardar imagen adjunta (si viene en el POST)
            # El formulario ya acepta request.FILES.
            # Si no se sube nada, se mantiene en None.
            if 'image' in request.FILES:
                form_request.image = request.FILES.get('image')
            elif 'images' in request.FILES:
                # Compatibilidad por si el input usa nombre plural
                # (en caso de que luego queramos múltiples).
                uploaded = request.FILES.get('images')
                if uploaded:
                    form_request.image = uploaded

            # Valores del tipo/servicio para que el modelo valide
            form_request.service_type = 'garantia'
            form_request.service_option = 'reparacion'
            form_request.priority = 'normal'
            form_request.status = 'pending'

            form_request.save()


            messages.success(
                request,
                "¡Solicitud de reparación enviada exitosamente! Te contactaremos en 24 horas.",
            )
            return redirect("services_catalog")
        else:
            # Mostrar errores reales para depuración (no ocultarlos)
            messages.error(request, "Por favor corrija los errores en el formulario.")

    else:
        form = FormRequestForm(initial={
            'service_type': 'garantia',
            'service_option': 'reparacion',
            'priority': 'normal',
            'customer_name': request.user.get_full_name() or request.user.username,
            'email': request.user.email or '',
        })

    return render(
        request,
        "products/reparacion_form.html",
        {
            "form": form,
            "cart": cart,
            "cart_count": cart.item_count,
        },
    )




def mantenimiento_form(request):
    """Catálogo de mantenimiento.

    Botón "Agregar al carrito": agrega un Service al carrito.
    Funciona igual que los demás productos: usa CartItem con el `service_id` real.
    """
    from cart.models import Cart, CartItem
    from django.contrib.contenttypes.models import ContentType

    cart = Cart.get_cart(request)
    cart_count = cart.item_count

    def add_service_to_cart(service_key: str):
        # Mapa clave -> lista de posibles valores dentro de `Service.name`.
        mapping = {
            "celular": ["celular", "telefono", "tel", "cel"],
            "laptop": ["laptop", "notebook", "portatil"],
            "tablet": ["tablet", "tab"],
            "audifonos": ["audif", "audífonos", "headphone", "earphone"],
            "revision_tecnica": ["revision", "revisi", "tecnica", "técnica"],
            "limpieza_profunda": ["limpieza", "profunda", "deep"],
        }

        candidates = mapping.get(service_key, ["mantenimiento"])

        q = Q()
        for c in candidates:
            q |= Q(name__icontains=c)

        service_qs = Service.objects.filter(q)
        if not service_qs.exists():
            service_qs = Service.objects.filter(name__icontains="mantenimiento")

        if not service_qs.exists():
            messages.warning(
                request,
                "No se encontró el servicio en el catálogo. Intenta con otro tipo.",
            )
            return

        service = service_qs.first()
        content_type = ContentType.objects.get_for_model(Service)

        cart_item, created = CartItem.objects.get_or_create(
            cart=cart,
            content_type=content_type,
            object_id=service.id,
            defaults={"quantity": 1},
        )
        if not created:
            cart_item.quantity += 1
            cart_item.save()

        messages.success(request, f"{service.name} agregado al carrito")

    # Botón "Agregar al carrito" (GET)
    if request.method == "GET" and request.GET.get("add_to_cart") == "on":
        service_key = (request.GET.get("service_key") or "").strip()

        # Asegura que el carrito exista en sesión (para que el redirect muestre el mismo carrito)
        _ = cart  # Cart.get_cart ya garantiza la creación

        add_service_to_cart(service_key)
        return redirect("cart_detail")

    # Mantener compatibilidad si alguien hace POST (no mostramos formulario en la plantilla)
    if request.method == "POST":
        messages.warning(
            request,
            "Este catálogo no tiene formulario. Usa el botón para agregar al carrito.",
        )
        return redirect("cart_detail")

    form = FormRequestForm()
    return render(
        request,
        "products/mantenimiento_form.html",
        {"form": form, "cart": cart, "cart_count": cart_count},
    )




def warranty_request(request):
    """Formulario de garantía/reparación para clientes"""
    from cart.models import Cart

    cart = Cart.get_cart(request)

    def _get_first_seller_profile():
        return UserProfile.objects.filter(role="vendedor").order_by("id").first()

    if request.method == "POST":
        form = FormRequestForm(request.POST, request.FILES)
        if form.is_valid():
            form_request = form.save(commit=False)

            # Datos del cliente
            form_request.customer_name = (
                request.user.get_full_name() or request.user.username
            )
            form_request.email = request.user.email or ""

            # Consistencia de tipo/servicio y estado
            form_request.service_type = "garantia"
            # El panel del vendedor filtra por reparacion/mantenimiento,
            # así que para que se vea aquí asignamos reparacion.
            if not form_request.service_option:
                form_request.service_option = "reparacion"

            form_request.priority = form_request.priority or "normal"
            form_request.status = form_request.status or "pending"

            # Asociar a vendedor para que aparezca en /productos/solicitudes/
            if not form_request.seller_id:
                form_request.seller = _get_first_seller_profile()

            form_request.save()
            messages.success(
                request,
                "Tu solicitud de garantia se ha enviado correctamente",
            )
            return redirect("home")
    else:
        form = FormRequestForm(initial={
            "service_type": "garantia",
            "service_option": "reparacion",
            "priority": "normal",
        })

    return render(request, "products/warranty_form.html", {"form": form, "cart": cart})



@login_required
def warranty_request_list(request):
    """Lista de solicitudes de garantía del cliente"""
    form_requests = FormRequest.objects.filter(
        models.Q(seller=request.user.profile) | models.Q(status="pending")
    ).order_by("-created_at")

    return render(
        request, "products/warranty_list.html", {"form_requests": form_requests}
    )


@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and getattr(u.profile, "role", None) == "vendedor")
def service_create(request):
    profile = request.user.profile

    if request.method == "POST":
        form = ServiceForm(request.POST)
        if form.is_valid():
            service = form.save(commit=False)
            service.seller = profile
            service.save()
            messages.success(request, "Servicio creado exitosamente.")
            return redirect("service_list")
    else:
        form = ServiceForm()
    return render(request, "products/service_form.html", {"form": form, "action": "Crear"})


@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and getattr(u.profile, "role", None) == "vendedor")
def service_list(request):
    profile = request.user.profile
    services = Service.objects.filter(seller=profile).order_by("-created_at")
    return render(request, "products/service_list.html", {"services": services})


@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and getattr(u.profile, "role", None) == "vendedor")
def service_update(request, pk):
    service = get_object_or_404(Service, pk=pk, seller=request.user.profile)
    if request.method == "POST":
        form = ServiceForm(request.POST, instance=service)
        if form.is_valid():
            form.save()
            messages.success(request, "Servicio actualizado exitosamente.")
            return redirect("service_list")
    else:
        form = ServiceForm(instance=service)
    return render(request, "products/service_form.html", {"form": form, "action": "Editar", "service": service})


@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and getattr(u.profile, "role", None) == "vendedor")
def service_delete(request, pk):
    service = get_object_or_404(Service, pk=pk, seller=request.user.profile)
    if request.method == "POST":
        service.delete()
        messages.success(request, "Servicio eliminado exitosamente.")
        return redirect("service_list")
    return render(request, "products/service_confirm_delete.html", {"service": service})


@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and getattr(u.profile, "role", None) == "vendedor")
def update_form_request_status(request, pk):
    """Actualiza el estado de una solicitud (AJAX) para el vendedor."""
    from django.http import JsonResponse

    if request.method != "POST":
        return JsonResponse({"ok": False, "message": "Método no permitido"}, status=405)

    profile = request.user.profile
    form_request = get_object_or_404(FormRequest, pk=pk, seller=profile)

    new_status = request.POST.get("status")
    allowed = {c[0] for c in FormRequest.STATUS_CHOICES}

    if not new_status:
        return JsonResponse({"ok": False, "message": "Estado no recibido"}, status=400)

    if new_status not in allowed:
        return JsonResponse({"ok": False, "message": "Estado inválido"}, status=400)

    try:
        form_request.status = new_status
        form_request.save(update_fields=["status", "updated_at"])
    except Exception as e:
        return JsonResponse(
            {
                "ok": False,
                "message": f"No se pudo guardar el estado: {str(e)}",
            },
            status=500,
        )

    return JsonResponse(
        {
            "ok": True,
            "status": form_request.status,
            "status_display": form_request.get_status_display(),
        }
    )


# Alias de compatibilidad con rutas existentes
vendor_update_form_request_status = update_form_request_status



@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and getattr(u.profile, "role", None) == "vendedor")
def vendor_download_form_request_image(request, pk):
    """Descarga la imagen adjunta desde el archivo almacenado en el modelo."""
    profile = request.user.profile
    form_request = get_object_or_404(FormRequest, pk=pk, seller=profile)

    if not form_request.image:
        return HttpResponse(status=404)

    # Importante: usar el archivo del ImageField (no copiar). 
    # FileResponse lee directamente desde el path real.
    return FileResponse(form_request.image.open('rb'), as_attachment=True, filename=form_request.image.name)


